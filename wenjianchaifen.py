import streamlit as st
import pandas as pd
import os
import zipfile
from io import BytesIO
import base64
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

# 初始化session state
if 'split_result' not in st.session_state:
    st.session_state.split_result = None
if 'merge_result' not in st.session_state:
    st.session_state.merge_result = None
if 'original_filename' not in st.session_state:
    st.session_state.original_filename = None
if 'uploaded_file_content' not in st.session_state:
    st.session_state.uploaded_file_content = None
if 'uploaded_files_content' not in st.session_state:
    st.session_state.uploaded_files_content = None

def get_excel_columns_openpyxl(file_content):
    """使用openpyxl读取Excel文件的列名而不加载数据"""
    # 以只读模式加载工作簿
    wb = load_workbook(file_content, read_only=True)
    ws = wb.active
    
    # 读取第一行作为列名
    columns = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()  # 关闭工作簿释放资源
    
    return columns

def split_excel_openpyxl(file_content, num_splits, selected_columns=None):
    """使用openpyxl拆分Excel文件（流式处理，低内存占用）"""
    wb = load_workbook(file_content, read_only=True)
    ws = wb.active
    
    # 获取总行列数
    total_rows = ws.max_row
    total_cols = ws.max_column
    
    # 转换列名索引（如"A", "B"）
    if selected_columns:
        # 将列名转换为索引（如"姓名"→"B"）
        col_indices = []
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        
        for col_name in selected_columns:
            if col_name in header:
                col_idx = header.index(col_name) + 1  # openpyxl从1开始计数
                col_indices.append(get_column_letter(col_idx))
        selected_cols = col_indices if col_indices else None
    else:
        selected_cols = None
    
    # 计算每个拆分文件的行数
    rows_per_split = total_rows // num_splits
    remainder = total_rows % num_splits
    split_plans = []
    current_row = 1  # openpyxl从1开始计数
    
    for i in range(num_splits):
        target_rows = rows_per_split + (1 if i < remainder else 0)
        split_plans.append((current_row, current_row + target_rows - 1))
        current_row += target_rows
    
    # 生成拆分文件数据
    split_dfs = []
    for start_row, end_row in split_plans:
        # 创建新的写入工作簿
        wb_split = load_workbook(write_only=True)
        ws_split = wb_split.create_sheet()
        
        # 写入表头
        header = [cell.value for cell in ws.iter_rows(min_row=1, max_row=1, max_col=total_cols if not selected_cols else len(selected_cols))]
        ws_split.append(header)
        
        # 流式写入数据行
        row_count = 0
        for row in ws.iter_rows(min_row=start_row, max_row=end_row):
            if selected_cols:
                # 只选择指定列
                values = [cell.value for cell in row if cell.column_letter in selected_cols]
            else:
                values = [cell.value for cell in row]
            ws_split.append(values)
            row_count += 1
        
        # 转换为DataFrame（仅用于预览，实际写入由openpyxl处理）
        if row_count > 0:
            df = pd.DataFrame(ws_split.values)
            df.columns = df.iloc[0]
            df = df[1:]
            split_dfs.append(df)
        
        wb_split.close()
    
    wb.close()
    return split_dfs

def merge_excel(files_content, selected_columns=None):
    """合并多个Excel文件为一个（一次性加载）"""
    dfs = []
    
    # 读取所有文件
    for file in files_content:
        df = pd.read_excel(file)
        
        # 筛选列
        if selected_columns:
            df = df[selected_columns]
            
        dfs.append(df)
    
    # 合并所有DataFrame
    merged_df = pd.concat(dfs, ignore_index=True)
    
    # 释放原始数据框内存
    del dfs
    
    return merged_df

def get_zip_download_link(split_dfs, original_filename):
    """生成包含所有拆分文件的ZIP下载链接"""
    zip_buffer = BytesIO()
    base_name, ext = os.path.splitext(original_filename)
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for i, df in enumerate(split_dfs):
            # 创建Excel文件
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
            excel_buffer.seek(0)
            
            # 添加到ZIP文件
            zipf.writestr(f"{base_name}——拆分{i+1}of{len(split_dfs)}{ext}", excel_buffer.getvalue())
    
    zip_buffer.seek(0)
    b64 = base64.b64encode(zip_buffer.read()).decode()
    href = f'<a href="data:application/zip;base64,{b64}" download="{base_name}_拆分文件.zip">下载所有拆分文件 (ZIP)</a>'
    
    # 释放内存
    del split_dfs
    
    return href

def get_excel_download_link(df, original_filename):
    """生成Excel下载链接"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='合并数据')
    output.seek(0)
    
    b64 = base64.b64encode(output.read()).decode()
    base_name, _ = os.path.splitext(original_filename if original_filename else "合并文件")
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{base_name}_合并.xlsx">下载合并后的Excel文件</a>'
    
    # 释放内存
    del df
    
    return href

def main():
    st.title("Excel文件拆分与合并工具（openpyxl优化版）")
    
    # 选择操作类型
    operation = st.radio("选择操作类型", ["拆分文件", "合并文件"])
    
    if operation == "拆分文件":
        st.header("Excel文件拆分（openpyxl优化）")
        
        # 上传文件并保存到session state
        uploaded_file = st.file_uploader("上传Excel文件", type=["xlsx", "xls"])
        if uploaded_file is not None:
            st.session_state.uploaded_file_content = uploaded_file
            st.session_state.original_filename = uploaded_file.name
        
        # 检查是否有上传的文件内容
        if st.session_state.uploaded_file_content is not None:
            # 读取列名
            with st.spinner("读取列名..."):
                all_columns = get_excel_columns_openpyxl(st.session_state.uploaded_file_content)
            
            # 设置拆分数量
            num_splits = st.number_input("拆分为几个文件", min_value=1, max_value=100, value=2)
            
            # 选择输出列
            selected_columns = st.multiselect(
                "选择要保留的列（未选择的列将被删除）",
                all_columns,
                default=all_columns
            )
            
            # 显示文件大小警告
            file_size = len(st.session_state.uploaded_file_content.getvalue()) / (1024 * 1024)  # MB
            if file_size > 100:
                st.warning(f"注意：文件大小为 {file_size:.2f} MB，openpyxl流式处理已优化内存占用。")
            
            if st.button("执行拆分"):
                if not selected_columns:
                    st.error("请至少选择一列")
                else:
                    # 重置结果
                    st.session_state.split_result = None
                    
                    # 读取数据并执行拆分
                    with st.spinner("正在使用openpyxl拆分文件..."):
                        try:
                            st.session_state.split_result = split_excel_openpyxl(
                                st.session_state.uploaded_file_content, 
                                num_splits, 
                                selected_columns
                            )
                            st.success(f"已成功将文件拆分为 {num_splits} 个部分")
                        except Exception as e:
                            st.error(f"处理过程中出错: {str(e)}")
                            st.error("请尝试减少拆分数量或选择更少的列，或使用更小的文件。")
            
            # 显示结果（如果有）
            if st.session_state.split_result is not None:
                st.success(f"已成功将文件拆分为 {num_splits} 个部分")
                
                # 显示前几个拆分文件的预览
                for i, split_df in enumerate(st.session_state.split_result[:3]):
                    st.subheader(f"拆分文件 {i+1}/{len(st.session_state.split_result)} 预览")
                    st.dataframe(split_df.head(10))
                
                # 生成ZIP下载链接
                st.markdown(get_zip_download_link(
                    st.session_state.split_result, 
                    st.session_state.original_filename
                ), unsafe_allow_html=True)
    
    else:  # 合并文件
        st.header("Excel文件合并")
        
        # 上传文件并保存到session state
        uploaded_files = st.file_uploader("上传多个Excel文件", type=["xlsx", "xls"], accept_multiple_files=True)
        if uploaded_files is not None and len(uploaded_files) > 0:
            st.session_state.uploaded_files_content = uploaded_files
            st.session_state.original_filename = uploaded_files[0].name if uploaded_files else "合并文件"
        
        # 检查是否有上传的文件内容
        if st.session_state.uploaded_files_content is not None and len(st.session_state.uploaded_files_content) > 0:
            # 读取第一个文件的列名
            with st.spinner("读取列名..."):
                all_columns = get_excel_columns_openpyxl(st.session_state.uploaded_files_content[0])
            
            # 选择输出列
            selected_columns = st.multiselect(
                "选择要保留的列（未选择的列将被删除）",
                all_columns,
                default=all_columns
            )
            
            # 计算总文件大小
            total_size = sum(len(file.getvalue()) for file in st.session_state.uploaded_files_content) / (1024 * 1024)
            if total_size > 100:
                st.warning(f"注意：文件总大小为 {total_size:.2f} MB，可能需要较长时间处理且占用较多内存。")
            
            if st.button("执行合并"):
                if not selected_columns:
                    st.error("请至少选择一列")
                elif not st.session_state.uploaded_files_content:
                    st.error("请上传至少一个文件")
                else:
                    # 重置结果
                    st.session_state.merge_result = None
                    
                    # 读取数据并执行合并
                    with st.spinner("正在读取数据并合并文件..."):
                        try:
                            st.session_state.merge_result = merge_excel(
                                st.session_state.uploaded_files_content, 
                                selected_columns
                            )
                            st.success(f"已成功合并 {len(st.session_state.uploaded_files_content)} 个文件")
                        except Exception as e:
                            st.error(f"合并过程中出错: {str(e)}")
                            st.error("可能是文件过大或格式不兼容，请尝试使用较小的文件或检查文件格式。")
            
            # 显示结果（如果有）
            if st.session_state.merge_result is not None:
                st.success(f"已成功合并 {len(st.session_state.uploaded_files_content)} 个文件")
                
                # 显示合并结果预览
                st.dataframe(st.session_state.merge_result.head(20))
                
                # 生成Excel下载链接
                st.markdown(get_excel_download_link(
                    st.session_state.merge_result, 
                    st.session_state.original_filename
                ), unsafe_allow_html=True)

if __name__ == "__main__":
    main()
